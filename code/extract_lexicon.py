#!/usr/bin/env python3
"""Признаки по опубликованным словарям: L07 и L09.

Источники и определения зафиксированы амендментом от 2026-07-25
(`02-preregistration/amendments.md`) до первого расчёта. Файлы словарей лежат
в `06-features/lexicons/` вместе с записанными sha256 — скрипт проверяет хеш
на старте и отказывается считать при расхождении.

Запуск из корня папки исследования:
    python 09-tools/extract_lexicon.py
    python 09-tools/extract_lexicon.py --limit 20

L07 — средний индекс абстрактности существительных по шкале источника (1–5),
во втором поле строки доля покрытия словарём.
L09 — токенов оценочной лексики на 1000 слов, во втором поле доля негативных
среди найденных.

Признаки D01, D02, D06 и D07 здесь не считаются: источника нет, решение
записано тем же амендментом.
"""

import argparse
import csv
import gzip
import hashlib
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import retest_io

sys.path.insert(0, str(Path(__file__).resolve().parent))
import feature_cache as fc  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent

for stream in (sys.stdout, sys.stderr):
    if hasattr(stream, "reconfigure"):
        stream.reconfigure(encoding="utf-8", errors="replace")

DOCUMENTS = ROOT / "04-corpus" / "documents-registry.csv"
STANZA_CACHE = ROOT / "06-features" / "cache" / "stanza-v1"
LEXICONS = ROOT / "06-features" / "lexicons"
MATRIX = ROOT / "06-features" / "feature-matrix.csv"
SCHEMA = ROOT / "06-features" / "feature-matrix-schema.csv"

EXTRACTOR_VERSION = "lex-v1"
# prep-v5, 2026-07-29: версия препроцессинга задаётся параметром, значение по
# умолчанию не меняется.
PREP_VERSION = "prep-v4"
STANZA_REVISION = "stanza 1.14.0/ru-syntagrus"
OWNED = ("L07", "L09")

# Хеши записаны в амендменте от 2026-07-25. Расхождение означает, что словарь
# подменили после фиксации, и считать по нему нельзя.
EXPECTED = {
    "kfu-abstractness-bert-22k.xlsx": "830617feda0f8dfcb585054b1b488b194942c01a0d123d355dce05cce22e9d8b",
    "rusentilex-2017.txt": "7d50cca27cf096098cde5d25cede66374b521b4769efbc46633fdd8b2c627b1a",
}

PUNCT_POS = "PUNCT"
MIN_WORDS = 100  # порог кодбука, раздел 10


def read_rows(path):
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def verify_lexicons():
    for name, expected in EXPECTED.items():
        path = LEXICONS / name
        if not path.exists():
            raise SystemExit(f"нет файла словаря: {path}")
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        if digest != expected:
            raise SystemExit(f"{name}: sha256 {digest} не совпадает с зафиксированным {expected}")
    print("словари проверены по sha256")


def load_abstractness():
    from openpyxl import load_workbook

    workbook = load_workbook(LEXICONS / "kfu-abstractness-bert-22k.xlsx", read_only=True)
    sheet = workbook.active
    scores = {}
    for word, score in sheet.iter_rows(values_only=True):
        if word is None or str(word).strip() == "слово":
            continue
        try:
            scores[str(word).strip().lower()] = float(score)
        except (TypeError, ValueError):
            continue
    workbook.close()
    return scores


def load_sentiment():
    """Лемма -> множество тональностей. Формат описан в шапке файла."""
    entries = {}
    with (LEXICONS / "rusentilex-2017.txt").open(encoding="utf-8") as fh:
        for line in fh:
            if line.startswith("!") or not line.strip():
                continue
            parts = [part.strip() for part in line.split(",")]
            if len(parts) < 4:
                continue
            lemma, sentiment = parts[2].lower(), parts[3]
            entries.setdefault(lemma, set()).add(sentiment)
    return entries


def document_values(parsed, abstractness, sentiment):
    values = {}
    skipped = {}

    words = [token for sentence in parsed["sentences"] for token in sentence if token["p"] != PUNCT_POS]
    if len(words) < MIN_WORDS:
        return {}, {"L07": "меньше 100 слов", "L09": "меньше 100 слов"}

    nouns = [token["l"].lower() for token in words if token["p"] == "NOUN"]
    found = [abstractness[lemma] for lemma in nouns if lemma in abstractness]
    if not found:
        skipped["L07"] = "ни одно существительное не найдено в словаре"
    else:
        values["L07"] = (
            "Абстрактность существительных, среднее",
            sum(found) / len(found),
            len(found) / len(nouns),
            "индекс 1–5 / доля покрытия",
        )

    hits = [sentiment[token["l"].lower()] for token in words if token["l"].lower() in sentiment]
    per_thousand = 1000.0 * len(hits) / len(words)
    negative = sum(1 for tags in hits if tags == {"negative"})
    values["L09"] = (
        "Оценочная лексика",
        per_thousand,
        (negative / len(hits)) if hits else 0.0,
        "на 1000 слов / доля негативных",
    )
    return values, skipped


def make_record(doc_id, feature_id, name, raw, normalized, measure, computed_at, reason=""):
    return {
        "document_id": doc_id,
        "feature_id": feature_id,
        "feature_name": name,
        "extractor_version": EXTRACTOR_VERSION,
        "preprocessing_profile": "prose",
        "raw_value": "" if raw is None else f"{raw:.6g}",
        "normalized_value": "" if normalized is None else f"{normalized:.6g}",
        "unit": measure,
        "genre_percentile": "",
        "missing_reason": reason,
        "computed_at": computed_at,
    }


def merge_into_matrix(records, registry):
    with SCHEMA.open(encoding="utf-8-sig") as fh:
        fields = next(csv.reader(fh))

    kept = []
    with MATRIX.open(encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            if row["document_id"] not in registry or row["feature_id"] in OWNED:
                continue
            kept.append(row)

    merged = kept + records

    if fc.percentiles_inline(PREP_VERSION):
        pools = {}
        for record in merged:
            if record["normalized_value"] == "" and record["raw_value"] == "":
                continue
            genre = registry[record["document_id"]]["genre"]
            value = float(record["normalized_value"] or record["raw_value"])
            pools.setdefault((record["feature_id"], genre), []).append(value)
        for key in pools:
            pools[key].sort()
        for record in merged:
            if record["normalized_value"] == "" and record["raw_value"] == "":
                continue
            genre = registry[record["document_id"]]["genre"]
            pool = pools[(record["feature_id"], genre)]
            value = float(record["normalized_value"] or record["raw_value"])
            rank = sum(1 for item in pool if item < value)
            record["genre_percentile"] = f"{rank / len(pool):.4f}" if pool else ""

    backup = MATRIX.with_suffix(".csv.bak-before-lex-v1")
    if not backup.exists():
        backup.write_bytes(MATRIX.read_bytes())
    with MATRIX.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(merged)
    print(f"матрица переписана: {MATRIX.relative_to(ROOT)}, строк {len(merged)}")


def main():
    global PREP_VERSION, MATRIX
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--ids-file", help="считать только документы из файла, по идентификатору на строку")
    parser.add_argument("--out", help="записать результат в CSV вместо слияния в матрицу")
    parser.add_argument("--prep-version", default=PREP_VERSION,
                        help="версия препроцессинга на входе")
    args = parser.parse_args()

    PREP_VERSION = args.prep_version
    MATRIX = fc.matrix_path(PREP_VERSION, MATRIX)

    verify_lexicons()
    abstractness = load_abstractness()
    sentiment = load_sentiment()
    print(f"словарь абстрактности: {len(abstractness)} лемм")
    print(f"RuSentiLex: {len(sentiment)} лемм")

    rows = read_rows(DOCUMENTS)
    registry = {row["document_id"]: row for row in rows}
    if args.ids_file:
        wanted = set(retest_io.read_ids(args.ids_file))
        rows = [row for row in rows if row["document_id"] in wanted]
    if args.limit:
        rows = rows[: args.limit]

    computed_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    records = []
    missing_cache = []
    coverage = []
    # Разбор берётся по ключу входа: у 68 документов prep-v5 текст изменился при
    # прежнем document_id, и адресация одним именем дала бы чужой разбор.
    index = fc.load_index(STANZA_CACHE)

    for row in rows:
        doc_id = row["document_id"]
        input_sha = fc.sha_for(PREP_VERSION, "prose", doc_id)
        path = (fc.lookup(STANZA_CACHE, index, doc_id, input_sha, STANZA_REVISION)
                if input_sha else None)
        if path is None:
            missing_cache.append(doc_id)
            continue
        with gzip.open(path, "rt", encoding="utf-8") as fh:
            parsed = json.load(fh)
        values, skipped = document_values(parsed, abstractness, sentiment)
        for feature_id, (name, raw, normalized, measure) in values.items():
            records.append(make_record(doc_id, feature_id, name, raw, normalized, measure, computed_at))
            if feature_id == "L07":
                coverage.append(normalized)
        for feature_id, reason in skipped.items():
            records.append(make_record(doc_id, feature_id, "", None, None, "", computed_at, reason))

    # Проба на части корпуса матрицу не трогает: усечённый прогон заменил бы
    # строки всех документов строками нескольких.
    if args.out:
        written = retest_io.write_records(args.out, records)
        print(f"повторный прогон на {len(rows)} документах, матрица не изменена: {args.out}, строк {written}")
    elif args.limit:
        print(f"проба на {len(rows)} документах, матрица не изменена")
    else:
        merge_into_matrix(records, registry)

    computed = sum(1 for record in records if not record["missing_reason"])
    print(f"L07 и L09 посчитаны: строк {len(records)}, со значением {computed}")
    if coverage:
        coverage.sort()
        print(f"  покрытие словарём абстрактности: медиана {coverage[len(coverage) // 2]:.3f}, минимум {coverage[0]:.3f}")
    if missing_cache:
        print(f"  ! нет разбора Stanza: {len(missing_cache)} документов")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
